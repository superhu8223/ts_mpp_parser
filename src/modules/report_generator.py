#!/usr/bin/env python
# -*- coding: utf-8 -*-
"""
报告生成模块
生成测试case和task的Markdown报告、Excel报表，以及统计图图片
"""

import os
from datetime import datetime
from typing import List
import matplotlib
matplotlib.use('Agg')  # 后端使用Agg，便于无界面生成图片
from matplotlib.figure import Figure
from matplotlib.backends.backend_agg import FigureCanvasAgg as FigureCanvas
import matplotlib.dates as mdates
import matplotlib.cm as cm
from matplotlib.colors import to_hex
try:
    import openpyxl
    from openpyxl.styles import Font, Alignment, PatternFill
    EXCEL_AVAILABLE = True
except ImportError:
    EXCEL_AVAILABLE = False
    print("警告: 未安装openpyxl库，Excel导出功能不可用。请运行: pip install openpyxl")
from .models import TestCase, TestTask, RuntimeEvent, StreamStats, EventLevel


class ReportGenerator:
    """报告生成器"""
    
    @staticmethod
    def generate_case_report(test_case: TestCase, stream_stats_list: List[StreamStats],
                            free_mem_history: List, case_dir: str):
        """生成测试case报告"""
        # 调试信息：打印stream_stats_list的详情
        print(f"[报告生成] generate_case_report被调用，stream_stats_list长度={len(stream_stats_list)}")
        for idx, stats in enumerate(stream_stats_list):
            print(f"[报告生成] stream_stats_list[{idx}]: stream_id={stats.stream_id}, url={stats.url}")
        
        report_path = os.path.join(case_dir, "case_report.md")
        
        with open(report_path, 'w', encoding='utf-8') as f:
            f.write(f"# 测试Case报告\n\n")
            f.write(f"## 基本信息\n\n")
            f.write(f"- **Case ID**: {test_case.case_id}\n")
            f.write(f"- **Case Name**: {test_case.case_name}\n")
            f.write(f"- **Run Command**: `{test_case.run_cmd}`\n")
            f.write(f"- **Stream Number**: {test_case.stream_num}\n")
            
            # bNeedCtrlC模式的特殊信息
            if test_case.b_need_ctrl_c:
                f.write(f"- **测试模式**: Ctrl+C循环模式（周期性退出重启）\n")
                f.write(f"- **Ctrl+C发送次数**: {test_case.ctrl_c_count}\n")
                f.write(f"- **循环周期**: {test_case.hold_time}秒\n")
            
            f.write(f"- **Start Time**: {test_case.start_time.strftime('%Y-%m-%d %H:%M:%S') if test_case.start_time else 'N/A'}\n")
            f.write(f"- **End Time**: {test_case.end_time.strftime('%Y-%m-%d %H:%M:%S') if test_case.end_time else 'N/A'}\n")
            
            if test_case.start_time and test_case.end_time:
                duration = (test_case.end_time - test_case.start_time).total_seconds()
                f.write(f"- **Duration**: {duration:.2f}秒\n")
            
            f.write(f"\n## Runtime事件\n\n")
            
            # bNeedCtrlC模式：汇总error和fatal事件
            if test_case.b_need_ctrl_c:
                error_events = [e for e in test_case.event_list if e.event_level == EventLevel.ERROR]
                fatal_events = [e for e in test_case.event_list if e.event_level == EventLevel.FATAL]
                
                f.write(f"### 错误与致命事件汇总\n\n")
                f.write(f"- **ERROR级别事件**: {len(error_events)}个\n")
                f.write(f"- **FATAL级别事件**: {len(fatal_events)}个\n\n")
                
                if error_events:
                    f.write(f"#### ERROR级别事件\n\n")
                    f.write(f"| 时间 | 简称 | 行日志 |\n")
                    f.write(f"|------|------|--------|\n")
                    for event in error_events:
                        clean_log = ReportGenerator._strip_device_timestamp(event.line_log)
                        f.write(f"| {event.timestamp.strftime('%H:%M:%S.%f')[:-3]} | "
                              f"{event.event_name} | {clean_log} |\n")
                
                if fatal_events:
                    f.write(f"\n#### FATAL级别事件\n\n")
                    f.write(f"| 时间 | 简称 | 行日志 |\n")
                    f.write(f"|------|------|--------|\n")
                    for event in fatal_events:
                        clean_log = ReportGenerator._strip_device_timestamp(event.line_log)
                        f.write(f"| {event.timestamp.strftime('%H:%M:%S.%f')[:-3]} | "
                              f"{event.event_name} | {clean_log} |\n")
                
                f.write(f"\n### 所有事件详细列表\n\n")
            
            if test_case.event_list:
                f.write(f"| 时间 | 级别 | 简称 | 行日志 |\n")
                f.write(f"|------|------|------|--------|\n")
                for event in test_case.event_list:
                    clean_log = ReportGenerator._strip_device_timestamp(event.line_log)
                    f.write(f"| {event.timestamp.strftime('%H:%M:%S.%f')[:-3]} | "
                          f"{event.event_level.value} | {event.event_name} | "
                          f"{clean_log} |\n")
            else:
                f.write("无事件记录\n")
            
            # bNeedCtrlC模式不展示详细的RTSP统计（简化模式无统计数据）
            if not test_case.b_need_ctrl_c:
                f.write(f"\n## RTSP流统计\n\n")
                for stats in stream_stats_list:
                    f.write(f"### Stream {stats.stream_id}: {stats.url}\n\n")
                    f.write(f"- **第一帧时间**: {stats.first_frame_time.strftime('%H:%M:%S.%f')[:-3] if stats.first_frame_time else 'N/A'}\n")
                    
                    if stats.fps_history:
                        f.write(f"\n#### 帧率历史\n\n")
                        f.write(f"| 时间 | 帧率(fps) |\n")
                        f.write(f"|------|----------|\n")
                        for data in stats.fps_history[-10:]:  # 最后10条记录
                            f.write(f"| {data.timestamp.strftime('%H:%M:%S')} | {data.value:.2f} |\n")
                    
                    if stats.bitrate_history:
                        f.write(f"\n#### 码率历史\n\n")
                        f.write(f"| 时间 | 码率(kbps) |\n")
                        f.write(f"|------|----------|\n")
                        for data in stats.bitrate_history[-10:]:  # 最后10条记录
                            f.write(f"| {data.timestamp.strftime('%H:%M:%S')} | {data.value:.2f} |\n")
            
            f.write(f"\n## 文件列表\n\n")
            f.write(f"- 串口日志: [serial.log](serial.log)\n")
            f.write(f"- Launch线程日志: [launchThread.log](launchThread.log)\n")
            
            # bNeedCtrlC模式下没有monitor线程
            if not test_case.b_need_ctrl_c:
                f.write(f"- Monitor线程日志: [monitorThread.log](monitorThread.log)\n")
                f.write(f"- ISP日志: [ispLog.txt](ispLog.txt)\n")
            
            # 列出关键帧截图（bNeedCtrlC模式下没有截图）
            if not test_case.b_need_ctrl_c:
                for stats in stream_stats_list:
                    for i in range(1, 4):
                        img_file = f"stream{stats.stream_id}_keyframe{i}.jpg"
                        if os.path.exists(os.path.join(case_dir, img_file)):
                            f.write(f"- Stream {stats.stream_id} 关键帧{i}: [{img_file}]({img_file})\n")
        
        print(f"Case报告已生成: {report_path}")
        
        # 生成Excel报表（无论是否为 bNeedCtrlC 模式，都生成以记录其他数据）
        if EXCEL_AVAILABLE:
            ReportGenerator._generate_case_excel(test_case, stream_stats_list, free_mem_history, case_dir)

        # bNeedCtrlC模式下没有解码和统计数据，不生成图表
        if not test_case.b_need_ctrl_c:
            # 生成统计图图片
            ReportGenerator._generate_case_chart(test_case, stream_stats_list, free_mem_history, case_dir)
    
    @staticmethod
    def generate_task_report(test_task: TestTask, task_dir: str):
        """生成测试task报告"""
        report_path = os.path.join(task_dir, "task_report.md")
        
        with open(report_path, 'w', encoding='utf-8') as f:
            f.write(f"# 测试Task报告\n\n")
            f.write(f"## 基本信息\n\n")
            f.write(f"- **Task Name**: {test_task.task_name}\n")
            f.write(f"- **EVB IP**: {test_task.evb_ip}\n")
            f.write(f"- **Serial Port**: {test_task.serial_config.port}\n")
            f.write(f"- **Start Time**: {test_task.start_time.strftime('%Y-%m-%d %H:%M:%S') if test_task.start_time else 'N/A'}\n")
            f.write(f"- **End Time**: {test_task.end_time.strftime('%Y-%m-%d %H:%M:%S') if test_task.end_time else 'N/A'}\n")
            
            if test_task.start_time and test_task.end_time:
                duration = (test_task.end_time - test_task.start_time).total_seconds()
                f.write(f"- **Total Duration**: {duration:.2f}秒\n")
            
            f.write(f"\n## 测试Case列表\n\n")
            f.write(f"| Case ID | Case Name | 状态 | 失败原因 |\n")
            f.write(f"|---------|-----------|------|----------|\n")
            
            for case in test_task.case_list:
                # 判断case是否失败（有emerge/fatal级别事件）
                has_fatal = any(e.event_level.value in ['emerge', 'fatal'] for e in case.event_list)
                status = "失败" if has_fatal else "成功"
                
                fail_reason = ""
                if has_fatal:
                    fatal_events = [e for e in case.event_list if e.event_level.value in ['emerge', 'fatal']]
                    fail_reason = ", ".join([e.event_name for e in fatal_events[:3]])
                
                case_dir_name = f"case_{case.case_id}"
                f.write(f"| {case.case_id} | [{case.case_name}]({case_dir_name}/case_report.md) | "
                       f"{status} | {fail_reason} |\n")
            
            f.write(f"\n## 总结\n\n")
            total_cases = len(test_task.case_list)
            failed_cases = sum(1 for case in test_task.case_list 
                             if any(e.event_level.value in ['emerge', 'fatal'] for e in case.event_list))
            passed_cases = total_cases - failed_cases
            
            f.write(f"- **总计Case数**: {total_cases}\n")
            f.write(f"- **通过**: {passed_cases}\n")
            f.write(f"- **失败**: {failed_cases}\n")
            f.write(f"- **通过率**: {(passed_cases/total_cases*100 if total_cases > 0 else 0):.1f}%\n")
        
        print(f"Task报告已生成: {report_path}")
    
    @staticmethod
    def _generate_case_excel(test_case: TestCase, stream_stats_list: List[StreamStats],
                            free_mem_history: List, case_dir: str):
        """生成Case Excel报表"""
        excel_path = os.path.join(case_dir, "case_statistics.xlsx")
        
        try:
            wb = openpyxl.Workbook()
            
            # 删除默认的Sheet
            if 'Sheet' in wb.sheetnames:
                wb.remove(wb['Sheet'])
            
            # Sheet1: Runtime事件清单
            ws_events = wb.create_sheet("Runtime事件清单")
            ws_events.append(["序号", "时间", "级别", "事件简称", "行日志"])
            
            # 设置标题行样式
            header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
            header_font = Font(bold=True, color="FFFFFF")
            for cell in ws_events[1]:
                cell.fill = header_fill
                cell.font = header_font
                cell.alignment = Alignment(horizontal="center", vertical="center")
            
            for idx, event in enumerate(test_case.event_list, 1):
                ws_events.append([
                    idx,
                    event.timestamp.strftime('%Y-%m-%d %H:%M:%S.%f')[:-3] if event.timestamp else '',
                    event.event_level.value,
                    event.event_name,
                    ReportGenerator._strip_device_timestamp(event.line_log)
                ])
            
            # 调整列宽
            ws_events.column_dimensions['A'].width = 8
            ws_events.column_dimensions['B'].width = 22
            ws_events.column_dimensions['C'].width = 12
            ws_events.column_dimensions['D'].width = 20
            ws_events.column_dimensions['E'].width = 80
            
            # Sheet2: 统计帧率
            ws_fps = wb.create_sheet("统计帧率")
            ws_fps.append(["Stream ID", "URL", "时间", "帧率(fps)"])
            for cell in ws_fps[1]:
                cell.fill = header_fill
                cell.font = header_font
                cell.alignment = Alignment(horizontal="center", vertical="center")
            
            for stats in stream_stats_list:
                for data in stats.fps_history:
                    ws_fps.append([
                        stats.stream_id,
                        stats.url,
                        data.timestamp.strftime('%Y-%m-%d %H:%M:%S') if data.timestamp else '',
                        round(data.value, 2)
                    ])
            
            ws_fps.column_dimensions['A'].width = 12
            ws_fps.column_dimensions['B'].width = 40
            ws_fps.column_dimensions['C'].width = 20
            ws_fps.column_dimensions['D'].width = 15
            
            # Sheet3: 统计码率
            ws_bitrate = wb.create_sheet("统计码率")
            ws_bitrate.append(["Stream ID", "URL", "时间", "码率(kbps)"])
            for cell in ws_bitrate[1]:
                cell.fill = header_fill
                cell.font = header_font
                cell.alignment = Alignment(horizontal="center", vertical="center")
            
            for stats in stream_stats_list:
                for data in stats.bitrate_history:
                    ws_bitrate.append([
                        stats.stream_id,
                        stats.url,
                        data.timestamp.strftime('%Y-%m-%d %H:%M:%S') if data.timestamp else '',
                        round(data.value, 2)
                    ])
            
            ws_bitrate.column_dimensions['A'].width = 12
            ws_bitrate.column_dimensions['B'].width = 40
            ws_bitrate.column_dimensions['C'].width = 20
            ws_bitrate.column_dimensions['D'].width = 15
            
            # Sheet4: free空闲内存
            ws_mem = wb.create_sheet("free空闲内存")
            ws_mem.append(["时间", "空闲内存(KB)", "总内存(KB)", "使用率(%)"])
            for cell in ws_mem[1]:
                cell.fill = header_fill
                cell.font = header_font
                cell.alignment = Alignment(horizontal="center", vertical="center")
            
            for mem_data in free_mem_history:
                if isinstance(mem_data, dict):
                    ws_mem.append([
                        mem_data.get('timestamp', ''),
                        mem_data.get('free', 0),
                        mem_data.get('total', 0),
                        round(mem_data.get('usage', 0), 2)
                    ])
            
            ws_mem.column_dimensions['A'].width = 20
            ws_mem.column_dimensions['B'].width = 18
            ws_mem.column_dimensions['C'].width = 18
            ws_mem.column_dimensions['D'].width = 15
            
            wb.save(excel_path)
            print(f"Excel报表已生成: {excel_path}")
            
        except Exception as e:
            print(f"生成Excel报表失败: {e}")

    @staticmethod
    def _generate_case_chart(test_case: TestCase, stream_stats_list: List[StreamStats],
                            free_mem_history: List, case_dir: str):
        """生成包含多路rtsp帧率/码率、free内存的统计图"""
        try:
            # 调试信息：检查stream_stats_list
            print(f"[报告图表] 开始绘制，共{len(stream_stats_list)}路stream")
            for stats in stream_stats_list:
                print(f"[报告图表] Stream{stats.stream_id}: bitrate_history={len(stats.bitrate_history) if stats.bitrate_history else 0}, fps_history={len(stats.fps_history) if stats.fps_history else 0}")
            
            # 如果没有任何数据，使用默认配置
            num_streams = len(stream_stats_list)
            if num_streams == 0:
                print("[报告图表] 警告：没有stream数据")
                return
            
            # 创建图表，使用 GridSpec 分割为3个子图
            from matplotlib.gridspec import GridSpec
            fig = Figure(figsize=(14, 10), dpi=100)
            canvas = FigureCanvas(fig)
            gs = GridSpec(3, 1, figure=fig, height_ratios=[2, 2, 1.2], hspace=0.4)
            
            ax_bitrate = fig.add_subplot(gs[0])
            ax_fps = fig.add_subplot(gs[1], sharex=ax_bitrate)
            ax_mem = fig.add_subplot(gs[2], sharex=ax_bitrate)

            # 颜色映射
            bitrate_cmap = cm.get_cmap('Greens')
            fps_cmap = cm.get_cmap('Blues')

            # 预先定义颜色（避免colormap索引问题）
            if num_streams > 1:
                bitrate_colors = [to_hex(bitrate_cmap(0.3 + 0.6 * i / (num_streams - 1))) for i in range(num_streams)]
                fps_colors = [to_hex(fps_cmap(0.3 + 0.6 * i / (num_streams - 1))) for i in range(num_streams)]
            else:
                bitrate_colors = [to_hex(bitrate_cmap(0.7))]
                fps_colors = [to_hex(fps_cmap(0.7))]
            
            print(f"[报告图表] 预生成颜色：bitrate_colors={bitrate_colors}, fps_colors={fps_colors}")

            # 配置码率轴
            ax_bitrate.set_ylabel('Bitrate (kbps)', color='green', fontsize=11, fontweight='bold')
            ax_bitrate.set_ylim(0, 6000)
            ax_bitrate.grid(True, alpha=0.2, linestyle='--')
            ax_bitrate.tick_params(axis='y', labelcolor='green')
            # 设置时间格式化
            ax_bitrate.xaxis.set_major_formatter(mdates.DateFormatter('%H:%M:%S'))

            # 配置帧率轴
            ax_fps.set_ylabel('FPS', color='blue', fontsize=11, fontweight='bold')
            ax_fps.set_ylim(0, 40)
            ax_fps.grid(True, alpha=0.2, linestyle='--')
            ax_fps.tick_params(axis='y', labelcolor='blue')

            # 配置内存轴
            ax_mem.set_ylabel('Free Mem (MB)', color='gray', fontsize=11, fontweight='bold')
            ax_mem.set_xlabel('Time', fontsize=11)
            ax_mem.set_ylim(0, 64)
            ax_mem.grid(True, alpha=0.2, linestyle='--')
            ax_mem.tick_params(axis='y', labelcolor='gray')

            # 时间格式
            ax_mem.xaxis.set_major_formatter(mdates.DateFormatter('%H:%M:%S'))
            fig.autofmt_xdate()

            # 绘制各路码率/帧率（连线，多路使用不同深浅的颜色）
            bitrate_plot_count = 0
            fps_plot_count = 0
            max_bitrate_value = 0.0  # 追踪所有stream的最大码率值
            
            for idx, stats in enumerate(stream_stats_list):
                print(f"[报告图表] 处理Stream{stats.stream_id}（列表索引{idx}）")
                
                if idx >= len(bitrate_colors):
                    print(f"[报告图表] 警告：索引{idx}超出颜色数组范围({len(bitrate_colors)})")
                    continue
                
                # 码率 - 连线，使用不同深浅的绿色
                if stats.bitrate_history and len(stats.bitrate_history) > 0:
                    print(f"[报告图表] 处理Stream{stats.stream_id}的bitrate_history，长度={len(stats.bitrate_history)}")
                    
                    # 详细检查每个数据点
                    times = []
                    vals = []
                    for i, d in enumerate(stats.bitrate_history):
                        if d.timestamp:
                            try:
                                # 与fps保持一致，直接使用mdates.date2num转换
                                time_num = mdates.date2num(d.timestamp)
                                times.append(time_num)
                                vals.append(d.value)
                            except Exception as e:
                                print(f"[报告图表] Stream{stats.stream_id}的第{i}个bitrate数据点转换失败: timestamp={d.timestamp}, error={e}")
                        else:
                            print(f"[报告图表] Stream{stats.stream_id}的第{i}个bitrate数据点timestamp为None")
                    
                    print(f"[报告图表] Stream{stats.stream_id}bitrate提取后: times={len(times)}, vals={len(vals)}")
                    if times and vals:
                        # 检查时间范围
                        time_diff = max(times) - min(times) if len(times) > 1 else 0
                        print(f"[报告图表] Stream{stats.stream_id}码率时间跨度: {time_diff:.1f}天，值范围: {min(vals):.1f}-{max(vals):.1f} kbps")
                    
                    if times and vals:  # 确保有有效数据
                        max_bitrate_value = max(max_bitrate_value, max(vals))  # 更新最大值
                        color = bitrate_colors[idx]
                        print(f"[报告图表] Stream{stats.stream_id}绘制{len(vals)}个码率数据点，时间范围={min(times):.6f}-{max(times):.6f}，颜色={color}，索引={idx}")
                        ax_bitrate.plot(times, vals, color=color, linewidth=2.5, marker='o', markersize=4,
                                       label=f'Stream{stats.stream_id}', linestyle='-', alpha=0.8)
                        bitrate_plot_count += 1
                    else:
                        print(f"[报告图表] Stream{stats.stream_id}没有有效的码率数据点")
                else:
                    print(f"[报告图表] Stream{stats.stream_id}没有码率历史数据")

                # 帧率 - 连线，使用不同深浅的蓝色
                if stats.fps_history and len(stats.fps_history) > 0:
                    print(f"[报告图表] 处理Stream{stats.stream_id}的fps_history，长度={len(stats.fps_history)}")
                    # 详细检查每个数据点
                    times = []
                    vals = []
                    for i, d in enumerate(stats.fps_history):
                        if d.timestamp:
                            try:
                                time_num = mdates.date2num(d.timestamp)
                                times.append(time_num)
                                vals.append(d.value)
                            except Exception as e:
                                print(f"[报告图表] Stream{stats.stream_id}的第{i}个fps数据点转换失败: timestamp={d.timestamp}, error={e}")
                        else:
                            print(f"[报告图表] Stream{stats.stream_id}的第{i}个fps数据点timestamp为None, value={d.value}")
                    
                    print(f"[报告图表] Stream{stats.stream_id}提取后: times={len(times)}, vals={len(vals)}")
                    if times and vals:  # 确保有有效数据
                        color = fps_colors[idx]
                        print(f"[报告图表] Stream{stats.stream_id}绘制{len(vals)}个帧率数据点，时间范围={min(times):.1f}-{max(times):.1f}，颜色={color}，索引={idx}")
                        ax_fps.plot(times, vals, color=color, linewidth=2.5, marker='s', markersize=4,
                                   label=f'Stream{stats.stream_id}', linestyle='-', alpha=0.8)
                        fps_plot_count += 1
                    else:
                        print(f"[报告图表] Stream{stats.stream_id}没有有效的帧率数据点（times={len(times)}, vals={len(vals)}）")
                else:
                    print(f"[报告图表] Stream{stats.stream_id}没有帧率历史数据或为空（fps_history={stats.fps_history}）")
            
            print(f"[报告图表] 绘图完成：绘制了{bitrate_plot_count}条码率线，{fps_plot_count}条帧率线")
            
            # 根据实际数据动态设置bitrate Y轴范围
            if max_bitrate_value > 0:
                ax_bitrate.set_ylim(0, max_bitrate_value * 1.1)
                print(f"[报告图表] 设置bitrate Y轴范围: 0 - {max_bitrate_value * 1.1:.1f} kbps (max={max_bitrate_value:.1f})")
            else:
                print(f"[报告图表] 未检测到有效bitrate数据，使用默认Y轴范围")

            # 绘制free内存（描点连线）
            if free_mem_history:
                print(f"[报告图表] free_mem_history有{len(free_mem_history)}条记录")
                mem_times = []
                mem_vals = []
                for m in free_mem_history:
                    ts = m.get('timestamp')
                    if isinstance(ts, str):
                        try:
                            ts = datetime.fromisoformat(ts)
                        except Exception:
                            ts = None
                    if ts:
                        mem_times.append(mdates.date2num(ts))
                        mem_vals.append(m.get('free', 0) / 1024)  # KB -> MB
                if mem_times:
                    print(f"[报告图表] 绘制{len(mem_times)}个内存数据点，范围: {min(mem_vals):.2f}MB - {max(mem_vals):.2f}MB")
                    ax_mem.plot(mem_times, mem_vals, color='gray', marker='^', markersize=5, 
                               linestyle='-', linewidth=2, label='Free Mem', alpha=0.7)
                    
                    # 在内存图上标记ERROR和FATAL事件
                    if test_case.event_list:
                        print(f"[报告图表] 标记{len(test_case.event_list)}个事件")
                        for event in test_case.event_list:
                            if event.event_level in [EventLevel.ERROR, EventLevel.FATAL]:
                                try:
                                    event_time = mdates.date2num(event.timestamp)
                                    if mem_times and mem_times[0] <= event_time <= mem_times[-1]:
                                        # 找到最接近的内存数据点
                                        idx = min(range(len(mem_times)), key=lambda i: abs(mem_times[i] - event_time))
                                        y_val = mem_vals[idx]
                                        marker_color = 'red' if event.event_level == EventLevel.FATAL else 'orange'
                                        marker_symbol = 'X' if event.event_level == EventLevel.FATAL else 'v'
                                        ax_mem.plot(event_time, y_val, marker=marker_symbol, markersize=12,
                                                   color=marker_color, markeredgecolor='black', markeredgewidth=1.5,
                                                   label=f'{event.event_level.value.upper()}: {event.event_name}',
                                                   zorder=10)
                                        print(f"[报告图表] 标记{event.event_level.value}事件: {event.event_name} at {event.timestamp.strftime('%H:%M:%S')}")
                                except Exception as marker_err:
                                    print(f"[报告图表] 标记事件失败: {marker_err}")
                    
                    ax_mem.legend(loc='upper right', fontsize=10)
                else:
                    print(f"[报告图表] 警告：没有有效的内存时间戳数据")
            else:
                print(f"[报告图表] 警告：free_mem_history为空")

            # 添加图例到码率和帧率轴
            if bitrate_plot_count > 0:
                ax_bitrate.legend(loc='upper left', fontsize=10, framealpha=0.95)
            if fps_plot_count > 0:
                ax_fps.legend(loc='upper left', fontsize=10, framealpha=0.95)

            fig.tight_layout()

            chart_path = os.path.join(case_dir, 'case_chart.png')
            fig.savefig(chart_path, dpi=100, bbox_inches='tight')
            fig.clear()
            print(f"统计图已生成: {chart_path}")
        except Exception as e:
            print(f"生成统计图失败: {e}")
            import traceback
            traceback.print_exc()

    @staticmethod
    def _strip_device_timestamp(line: str) -> str:
        """移除设备端时间戳/前缀，只保留日志主体。"""
        if not line:
            return ""
        import re
        # 匹配形如 "[1970-01-01 00:00:36.282486]", "1970_01_01 00:00:36.282486", "[    0.000000]" 等
        patterns = [
            r'^\s*\[\d{4}-\d{2}-\d{2}\s+\d{2}:\d{2}:\d{2}\.\d+\]\s*',  # [1970-01-01 HH:MM:SS.millisec]
            r'^\s*\d{4}_\d{2}_\d{2}\s+\d{2}:\d{2}:\d{2}\.\d+\s*',  # 1970_01_01 HH:MM:SS.millisec
            r'^\s*\[\s*\d+\.\d+\]\s*'  # [    0.000000]
        ]
        cleaned = line
        for pat in patterns:
            cleaned = re.sub(pat, '', cleaned)
        return cleaned.strip()
